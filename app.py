"""
Polla Mundial 2026 - Backend Flask
Lee las predicciones y resultados del Excel y calcula puntos.
"""

import os
import sys
import io
import threading
import webbrowser
from flask import Flask, jsonify, request
from flask_cors import CORS
import openpyxl


def get_base_path():
    """Ruta base de recursos. Funciona en desarrollo y en bundle PyInstaller."""
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS  # carpeta temporal del bundle
    return os.path.dirname(os.path.abspath(__file__))


def get_excel_path():
    """El Excel siempre vive junto al ejecutable (o al script en dev)."""
    if getattr(sys, 'frozen', False):
        return os.path.join(os.path.dirname(sys.executable), 'Polla Mundial 2026.xlsx')
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Polla Mundial 2026.xlsx')


BASE_PATH = get_base_path()
EXCEL_PATH = get_excel_path()

app = Flask(__name__, static_folder=os.path.join(BASE_PATH, 'static'))
CORS(app)


PARTICIPANTS = ['Hugo', 'Oscar', 'Camilo']

# Puntos por ronda de playoff
PLAYOFF_POINTS = {
    'Dieciseisavos': 4,
    'Octavos': 6,
    'Cuartos': 8,
    'Semifinal': 10,
    'Tercero': 12,
    'Final': 15,
    'Campeón': 20,
}

PLAYOFF_COLUMNS = ['Dieciseisavos', 'Octavos', 'Cuartos', 'Semifinal', 'Tercero', 'Final', 'Campeón']

# =====================================================
# Ranking FIFA de los 48 equipos participantes en el Mundial 2026
# Fuente: Rankings FIFA junio 2026 (oficiales pre-torneo)
# =====================================================
FIFA_RANKING = {
    'Argentina':            1,
    'España':               2,
    'Francia':              3,
    'Inglaterra':           4,
    'Portugal':             5,
    'Brasil':               6,
    'Marruecos':            7,
    'Países Bajos':         8,
    'Bélgica':              9,
    'Alemania':             10,
    'Croacia':              11,
    'Colombia':             13,
    'México':               14,
    'Senegal':              15,
    'Uruguay':              16,
    'Estados Unidos':       17,
    'Japón':                18,
    'Suiza':                19,
    'Irán':                 20,
    'Turquía':              22,
    'Ecuador':              23,
    'Australia':            27,
    'Costa de Marfil':      33,
    'Noruega':              34,
    'Escocia':              36,
    'Austria':              37,
    'República Checa':      40,
    'Paraguay':             41,
    'Canadá':               44,
    'Corea del Sur':        22,  # ~22nd
    'Suecia':               34,
    'Ghana':                50,
    'Argelia':              52,
    'Túnez':                36,
    'Egipto':               39,
    'Arabia Saudí':         56,
    'Sudáfrica':            60,
    'Jordania':             70,
    'RD Congo':             55,
    'Uzbekistán':           74,
    'Irak':                 67,
    'Bosnia y Herzegovina': 62,
    'Panamá':               80,
    'Haití':                83,
    'Cabo Verde':           90,
    'Catar':                57,
    'Curazao':              95,
    'Nueva Zelanda':        99,
}


def normalize_name(name):
    """Normaliza nombre de equipo para comparación insensible a tildes/acentos."""
    if not name:
        return ''
    replacements = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'Á': 'a', 'É': 'e', 'Í': 'i', 'Ó': 'o', 'Ú': 'u',
        'ü': 'u', 'Ü': 'u', 'ñ': 'n', 'Ñ': 'n',
    }
    result = str(name).strip().lower()
    for orig, repl in replacements.items():
        result = result.replace(orig, repl)
    return result


EXCEL_LOCKED_MSG = (
    'No se puede acceder al archivo Excel porque está abierto en otro programa '
    '(probablemente Microsoft Excel). Guarda los cambios, cierra el archivo y vuelve a intentar.'
)


def load_workbook(data_only=True):
    """Carga el workbook. Las lecturas usan memoria para no chocar con Excel abierto."""
    if data_only:
        try:
            with open(EXCEL_PATH, 'rb') as f:
                data = f.read()
        except PermissionError as exc:
            raise PermissionError(EXCEL_LOCKED_MSG) from exc
        return openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    try:
        return openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    except PermissionError as exc:
        raise PermissionError(EXCEL_LOCKED_MSG) from exc


def read_group_predictions(ws, max_row=73):
    """
    Lee los partidos de fase de grupos de una hoja de predicciones.
    Retorna lista de dicts con partido, grupo, local, visitante, g_local, g_visitante, resultado, diferencia.
    """
    matches = []
    for row in ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
        partido, grupo, local, visitante, g_local, g_visitante, resultado, diferencia = row[:8]
        if isinstance(partido, int):
            matches.append({
                'partido': partido,
                'grupo': grupo,
                'local': local,
                'visitante': visitante,
                'g_local': g_local,
                'g_visitante': g_visitante,
                'resultado': resultado,
                'diferencia': diferencia,
            })
    return matches


def read_playoffs(ws):
    """
    Lee la hoja de playoffs (Playoffs_XXX o Realidad Playoffs).
    Estructura: col A=Dieciseisavos, B=Octavos, C=Cuartos, D=Semifinal, E=Tercero, F=Final, G=Campeón
    Row 1 = headers, rows 2+ = equipos
    Retorna dict { 'Dieciseisavos': [equipo1, equipo2, ...], ... }
    """
    result = {ronda: [] for ronda in PLAYOFF_COLUMNS}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        for col_idx, ronda in enumerate(PLAYOFF_COLUMNS):
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                result[ronda].append(str(val).strip())
    return result


def read_playoff_matches(ws):
    """
    Lee una hoja de partidos de playoff (Realidad Playoffs Predicciones o Playoffs_Predicciones_*).
    Detecta bloques por título de sección ('PARTIDOS DE XXX') y retorna:
        { ronda_key: [ {partido, local, visitante, g_local, g_visitante, played}, ... ] }
    donde ronda_key es el valor de PLAYOFF_DISPLAY_ROUNDS que corresponda.
    """
    # Mapa de palabras clave en el título → clave de ronda
    SECTION_MAP = {
        'DIECISEISAVOS': 'Dieciseisavos',
        'OCTAVOS':       'Octavos',
        'CUARTOS':       'Cuartos',
        'SEMIFINAL':     'Semifinal',
        'TERCEROS':      'Tercero',
        'TERCERO':       'Tercero',
        'FINAL':         'Final',   # debe ir después de SEMIFINAL
    }
    if ws is None:
        return {}

    result = {}   # { ronda: [matches] }
    current_ronda = None

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        cell0 = str(row[0]).strip().upper()

        # ¿Es un encabezado de sección?
        matched_ronda = None
        for keyword, ronda in SECTION_MAP.items():
            if keyword in cell0:
                # Evitar que 'FINAL' capture 'SEMIFINAL'
                if keyword == 'FINAL' and 'SEMIFINAL' in cell0:
                    continue
                matched_ronda = ronda
                break
        if matched_ronda:
            current_ronda = matched_ronda
            if current_ronda not in result:
                result[current_ronda] = []
            continue

        # ¿Es fila de encabezado de columnas?
        if cell0 in ('PARTIDO', 'PARTIDOS'):
            continue

        # ¿Es fila de datos?
        if current_ronda is None:
            continue
        try:
            partido = int(row[0])
        except (TypeError, ValueError):
            continue

        local     = str(row[1]).strip() if row[1] is not None else ''
        visitante = str(row[2]).strip() if row[2] is not None else ''
        g_local     = int(row[3]) if row[3] is not None else None
        g_visitante = int(row[4]) if row[4] is not None else None
        played = g_local is not None and g_visitante is not None

        result[current_ronda].append({
            'partido':     partido,
            'local':       local,
            'visitante':   visitante,
            'g_local':     g_local,
            'g_visitante': g_visitante,
            'played':      played,
        })

    return result


def calculate_playoff_match_points(real, pred):
    """
    Calcula GL, GV, Resultado y Diferencia para un partido de playoff.
    real: dict con g_local, g_visitante, played
    pred: dict con g_local, g_visitante
    """
    zero = {'pts_g_local': 0, 'pts_g_visitante': 0, 'pts_resultado': 0, 'pts_diferencia': 0, 'total': 0}
    if not real['played'] or pred['g_local'] is None or pred['g_visitante'] is None:
        return zero

    rl, rv = int(real['g_local']), int(real['g_visitante'])
    pl, pv = int(pred['g_local']), int(pred['g_visitante'])

    pts_g_local     = 1 if rl == pl else 0
    pts_g_visitante = 1 if rv == pv else 0

    def resultado(gl, gv):
        return 'L' if gl > gv else ('E' if gl == gv else 'V')

    pts_resultado  = 1 if resultado(rl, rv) == resultado(pl, pv) else 0
    pts_diferencia = 1 if (rl - rv) == (pl - pv) else 0

    total = pts_g_local + pts_g_visitante + pts_resultado + pts_diferencia
    return {
        'pts_g_local':     pts_g_local,
        'pts_g_visitante': pts_g_visitante,
        'pts_resultado':   pts_resultado,
        'pts_diferencia':  pts_diferencia,
        'total':           total,
    }

def heal_real_playoffs(playoffs, matches_by_round):
    """
    Rellena los equipos faltantes en las siguientes rondas de playoffs usando los
    ganadores de los partidos de la ronda anterior, para compensar fórmulas vacías.
    También rellena local y visitante de los partidos.
    """
    def get_winner(match, next_round_teams):
        if not match.get('played'):
            return None
        gl, gv = match.get('g_local'), match.get('g_visitante')
        if gl is not None and gv is not None:
            if gl > gv: return match.get('local')
            if gv > gl: return match.get('visitante')
        
        local_n = normalize_name(match.get('local', ''))
        vis_n = normalize_name(match.get('visitante', ''))
        for t in next_round_teams:
            tn = normalize_name(t)
            if tn == local_n: return match.get('local')
            if tn == vis_n: return match.get('visitante')
        return None

    stages = [
        ('Dieciseisavos', 'Octavos'),
        ('Octavos', 'Cuartos'),
        ('Cuartos', 'Semifinal'),
        ('Semifinal', 'Final'),
    ]
    for prev_ronda, next_ronda in stages:
        prev_matches = matches_by_round.get(prev_ronda, [])
        next_matches = matches_by_round.get(next_ronda, [])
        
        # 1. Sanar local y visitante en los partidos de la siguiente ronda
        for i, m_next in enumerate(next_matches):
            p1_idx = i * 2
            p2_idx = i * 2 + 1
            if p1_idx < len(prev_matches) and not m_next.get('local'):
                w1 = get_winner(prev_matches[p1_idx], playoffs.get(next_ronda, []))
                if w1: m_next['local'] = w1
            if p2_idx < len(prev_matches) and not m_next.get('visitante'):
                w2 = get_winner(prev_matches[p2_idx], playoffs.get(next_ronda, []))
                if w2: m_next['visitante'] = w2

        # 2. Sanar equipos clasificados en playoffs
        for m in prev_matches:
            w = get_winner(m, playoffs.get(next_ronda, []))
            if w and normalize_name(w) not in [normalize_name(t) for t in playoffs.get(next_ronda, [])]:
                if not playoffs.get(next_ronda):
                    playoffs[next_ronda] = []
                playoffs[next_ronda].append(w)
                    
    if 'Final' in matches_by_round:
        for m in matches_by_round['Final']:
            w = get_winner(m, playoffs.get('Campeón', []))
            if w and normalize_name(w) not in [normalize_name(t) for t in playoffs.get('Campeón', [])]:
                if not playoffs.get('Campeón'):
                    playoffs['Campeón'] = []
                playoffs['Campeón'].append(w)

    return playoffs, matches_by_round


def calculate_group_standings(real_matches):
    """
    Calcula la tabla de posiciones de cada grupo a partir de los partidos jugados.
    Solo cuenta partidos con g_local y g_visitante no nulos.

    Retorna dict { grupo: [ {equipo, pts, gf, gc, dg, pj, pg, pe, pp}, ... ] }
    ordenado por: pts DESC, dg DESC, gf DESC, ranking_fifa ASC
    """
    # Recopilar datos por grupo
    groups = {}
    for m in real_matches:
        g = m['grupo']
        if g not in groups:
            groups[g] = {}
        for equipo in [m['local'], m['visitante']]:
            if equipo not in groups[g]:
                groups[g][equipo] = {'pts': 0, 'gf': 0, 'gc': 0, 'pj': 0, 'pg': 0, 'pe': 0, 'pp': 0}

    for m in real_matches:
        g = m['grupo']
        local = m['local']
        visitante = m['visitante']
        gl = m['g_local']
        gv = m['g_visitante']

        if gl is None or gv is None:
            continue

        gl, gv = int(gl), int(gv)
        groups[g][local]['pj'] += 1
        groups[g][visitante]['pj'] += 1
        groups[g][local]['gf'] += gl
        groups[g][local]['gc'] += gv
        groups[g][visitante]['gf'] += gv
        groups[g][visitante]['gc'] += gl

        if gl > gv:
            groups[g][local]['pts'] += 3
            groups[g][local]['pg'] += 1
            groups[g][visitante]['pp'] += 1
        elif gv > gl:
            groups[g][visitante]['pts'] += 3
            groups[g][visitante]['pg'] += 1
            groups[g][local]['pp'] += 1
        else:
            groups[g][local]['pts'] += 1
            groups[g][visitante]['pts'] += 1
            groups[g][local]['pe'] += 1
            groups[g][visitante]['pe'] += 1

    # Calcular diferencia de goles y ordenar
    result = {}
    for g, teams in groups.items():
        table = []
        for equipo, stats in teams.items():
            stats['equipo'] = equipo
            stats['dg'] = stats['gf'] - stats['gc']
            stats['fifa_rank'] = FIFA_RANKING.get(equipo, 999)
            table.append(stats)

        # Ordenar: pts DESC, dg DESC, gf DESC, fifa_rank ASC
        table.sort(key=lambda t: (-t['pts'], -t['dg'], -t['gf'], t['fifa_rank']))

        # Asignar posiciones
        for i, t in enumerate(table):
            t['pos'] = i + 1

        # Marcar si el grupo está completo (6 partidos = todas las combinaciones de 4 equipos)
        total_matches = sum(t['pj'] for t in table) // 2
        group_complete = total_matches == 6
        group_started = total_matches > 0

        result[g] = {
            'table': table,
            'complete': group_complete,
            'started': group_started,
            'matches_played': total_matches,
        }

    return result


def get_best_thirds(standings):
    """
    Obtiene los 8 mejores terceros de entre los 12 grupos.

    Criterios FIFA (en orden):
    1. Puntos
    2. Diferencia de gol
    3. Goles anotados
    4. Fair play (no disponible → omitimos)
    5. Ranking FIFA

    Solo se consideran grupos donde el 3er puesto esté determinado
    (al menos que hayan jugado suficientes partidos).
    Retorna: (mejores_8, todos_12, grupos_completos)
    """
    thirds = []
    for grupo, data in sorted(standings.items()):
        table = data['table']
        if len(table) >= 3:
            third = table[2].copy()
            third['grupo'] = grupo
            thirds.append(third)

    # Ordenar los 12 terceros
    thirds.sort(key=lambda t: (-t['pts'], -t['dg'], -t['gf'], t['fifa_rank']))

    # Los 8 mejores clasifican, los 4 restantes quedan eliminados
    best_8 = thirds[:8]
    rest_4 = thirds[8:]

    return best_8, rest_4, thirds


def get_qualified_teams(standings, best_8_thirds):
    """
    Retorna todos los equipos clasificados al dieciseisavos.
    - Top 2 de cada grupo (24 equipos)
    - 8 mejores terceros
    Total: 32 equipos
    """
    qualified = []
    thirds_set = {t['equipo'] for t in best_8_thirds}

    for grupo, data in sorted(standings.items()):
        table = data['table']
        for team in table:
            if team['pos'] == 1:
                qualified.append({'equipo': team['equipo'], 'grupo': grupo, 'tipo': '1°', 'pos': 1})
            elif team['pos'] == 2:
                qualified.append({'equipo': team['equipo'], 'grupo': grupo, 'tipo': '2°', 'pos': 2})
            elif team['pos'] == 3 and team['equipo'] in thirds_set:
                qualified.append({'equipo': team['equipo'], 'grupo': grupo, 'tipo': '3° ★', 'pos': 3})

    return qualified


def write_qualified_to_excel(qualified_teams, best_8_thirds):
    """
    Escribe los equipos clasificados en la hoja 'Realidad Playoffs' del Excel,
    columna A (Dieciseisavos).
    Solo escribe si hay equipos clasificados.
    """
    if not qualified_teams:
        return False

    # Cargar sin data_only para poder escribir
    wb = load_workbook(data_only=False)
    ws = wb['Realidad Playoffs']

    # Limpiar columna A (Dieciseisavos) desde fila 2
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].value = None

    # Escribir los equipos clasificados (32 equipos en col A)
    team_names = [t['equipo'] for t in qualified_teams]
    for i, name in enumerate(team_names):
        ws.cell(row=i + 2, column=1, value=name)

    try:
        wb.save(EXCEL_PATH)
    except PermissionError as exc:
        raise PermissionError(EXCEL_LOCKED_MSG) from exc
    return True


def _calc_resultado_diferencia(g_local, g_visitante):
    """Calcula resultado y diferencia a partir de los goles (evita depender de fórmulas Excel)."""
    if g_local is None or g_visitante is None:
        return None, None
    g_local, g_visitante = int(g_local), int(g_visitante)
    if g_local > g_visitante:
        return 'Local', g_local - g_visitante
    elif g_local < g_visitante:
        return 'Visitante', g_local - g_visitante
    else:
        return 'Empate', 0


def calculate_group_points(predictions, actuals):
    """
    Compara predicciones de grupos vs resultados reales.
    IMPORTANTE: Calcula resultado y diferencia desde los goles directamente,
    NO confía en las columnas G y H del Excel (son fórmulas → vienen como None).
    """
    actual_by_match = {m['partido']: m for m in actuals if m['partido'] is not None}
    details = []

    for pred in predictions:
        partido = pred['partido']
        real = actual_by_match.get(partido)

        pts_g_local = 0
        pts_g_visitante = 0
        pts_resultado = 0
        pts_diferencia = 0
        played = False

        if real and real.get('g_local') is not None and real.get('g_visitante') is not None:
            played = True
            r_g_local = int(real['g_local'])
            r_g_visitante = int(real['g_visitante'])
            # Calcular resultado/diferencia desde los goles (no confiar en columnas de fórmulas)
            r_resultado, r_diferencia = _calc_resultado_diferencia(r_g_local, r_g_visitante)

            p_g_local = pred.get('g_local')
            p_g_visitante = pred.get('g_visitante')
            # Para las predicciones también calculamos (son valores directos en el Excel)
            p_resultado, p_diferencia = _calc_resultado_diferencia(p_g_local, p_g_visitante)

            if p_g_local is not None and int(p_g_local) == r_g_local:
                pts_g_local = 1
            if p_g_visitante is not None and int(p_g_visitante) == r_g_visitante:
                pts_g_visitante = 1
            if p_resultado is not None and r_resultado is not None and p_resultado == r_resultado:
                pts_resultado = 1
            if p_diferencia is not None and r_diferencia is not None and p_diferencia == r_diferencia:
                pts_diferencia = 1

        details.append({
            'partido': partido,
            'grupo': pred['grupo'],
            'local': pred['local'],
            'visitante': pred['visitante'],
            'pred_g_local': pred.get('g_local'),
            'pred_g_visitante': pred.get('g_visitante'),
            'pred_resultado': pred.get('resultado'),
            'pred_diferencia': pred.get('diferencia'),
            'real_g_local': real.get('g_local') if real else None,
            'real_g_visitante': real.get('g_visitante') if real else None,
            'real_resultado': real.get('resultado') if real else None,
            'real_diferencia': real.get('diferencia') if real else None,
            'played': played,
            'pts_g_local': pts_g_local,
            'pts_g_visitante': pts_g_visitante,
            'pts_resultado': pts_resultado,
            'pts_diferencia': pts_diferencia,
            'total': pts_g_local + pts_g_visitante + pts_resultado + pts_diferencia,
        })
    return details


def calculate_playoff_points(pred_playoffs, real_playoffs):
    """
    Compara predicciones de playoffs vs resultados reales.
    Usa normalización de nombres para evitar errores por tildes.
    """
    details = {}
    total = 0

    for ronda in PLAYOFF_COLUMNS:
        puntos_ronda = PLAYOFF_POINTS[ronda]
        pred_raw = pred_playoffs.get(ronda, [])
        real_raw = real_playoffs.get(ronda, [])

        # Normalizar para comparación, manteniendo el nombre original
        pred_norm = {normalize_name(t): t for t in pred_raw if t}
        real_norm = {normalize_name(t): t for t in real_raw if t}

        acertados_norm = set(pred_norm.keys()) & set(real_norm.keys())
        # Recuperar nombres originales de los acertados (usar el de real)
        acertados = sorted([real_norm[k] for k in acertados_norm])

        pts = len(acertados) * puntos_ronda
        total += pts

        details[ronda] = {
            'predichos': sorted(pred_raw),
            'reales': sorted(real_raw),
            'acertados': acertados,
            'puntos_por_acierto': puntos_ronda,
            'total_ronda': pts,
        }

    return {'por_ronda': details, 'total': total}


@app.route('/')
def index():
    return app.send_static_file('index.html')


@app.route('/api/data')
def get_data():
    """
    Retorna todos los datos + standings de grupos + clasificación a playoffs.
    """
    try:
        wb = load_workbook()

        # Leer resultados reales de grupos
        ws_realidad = wb['Realidad']
        real_matches = read_group_predictions(ws_realidad)

        # Leer resultados reales de playoffs (equipos que avanzan)
        ws_real_playoffs = wb['Realidad Playoffs']
        real_playoffs = read_playoffs(ws_real_playoffs)

        # Leer resultados reales de partidos de playoff (match-level, por ronda)
        ws_real_pm = wb['Realidad Playoffs Predicciones'] if 'Realidad Playoffs Predicciones' in wb.sheetnames else None
        real_playoff_matches_by_round = read_playoff_matches(ws_real_pm)
        
        real_playoffs, real_playoff_matches_by_round = heal_real_playoffs(real_playoffs, real_playoff_matches_by_round)
        
        # Lista plana con campo 'ronda' (para backward-compat con el frontend)
        real_playoff_matches = [
            {**m, 'ronda': ronda}
            for ronda, matches in real_playoff_matches_by_round.items()
            for m in matches
        ]
        
        # Indexar por (ronda, partido)
        real_pm_index = {
            (m['ronda'], m['partido']): m
            for m in real_playoff_matches
        }

        # Calcular clasificaciones de grupos
        standings = calculate_group_standings(real_matches)
        best_8, rest_4, all_thirds = get_best_thirds(standings)
        qualified = get_qualified_teams(standings, best_8)

        # Determinar equipos eliminados para no contarlos como pendientes
        eliminated_from_tournament = set()
        eliminated_from_final = set()
        
        groups_complete = all(g['complete'] for g in standings.values()) if standings else False
        if groups_complete:
            qualified_names = {normalize_name(q['equipo']) for q in qualified}
            for g in standings.values():
                for t in g['table']:
                    t_n = normalize_name(t['equipo'])
                    if t_n not in qualified_names:
                        eliminated_from_tournament.add(t_n)
                        
        for m in real_playoff_matches:
            if m.get('played'):
                gl, gv = m.get('g_local'), m.get('g_visitante')
                if gl is not None and gv is not None:
                    loser = m.get('visitante') if gl > gv else m.get('local') if gv > gl else None
                    if loser:
                        loser_n = normalize_name(loser)
                        if m['ronda'] == 'Semifinal':
                            eliminated_from_final.add(loser_n)
                        else:
                            eliminated_from_tournament.add(loser_n)
                            
        def is_eliminated(team_n, r):
            if team_n in eliminated_from_tournament:
                return True
            if r in ('Final', 'Campeón') and team_n in eliminated_from_final:
                return True
            if r == 'Tercero' and team_n in [normalize_name(t) for t in real_playoffs.get('Final', [])]:
                return True
            return False

        # Leer predicciones de partidos de playoff por participante (también por ronda)
        playoff_match_preds = {}
        for name in PARTICIPANTS:
            ws_pm = wb[f'Playoffs_Predicciones_{name}'] if f'Playoffs_Predicciones_{name}' in wb.sheetnames else None
            pm_by_round = read_playoff_matches(ws_pm)
            # Indexar por (ronda, partido)
            playoff_match_preds[name] = {
                (ronda, m['partido']): m
                for ronda, matches in pm_by_round.items()
                for m in matches
            }

        # Ensamblar tabla combinada de partidos de playoff (para sección Grupos)
        playoff_all_matches = []
        for rm in real_playoff_matches:
            ronda = rm['ronda']
            pid   = rm['partido']
            entry = {
                'partido':          pid,
                'ronda':            ronda,
                'local':            rm['local'],
                'visitante':        rm['visitante'],
                'real_g_local':     rm['g_local'],
                'real_g_visitante': rm['g_visitante'],
                'played':           rm['played'],
                'participants': {},
            }
            for name in PARTICIPANTS:
                pred = playoff_match_preds[name].get((ronda, pid), {'g_local': None, 'g_visitante': None, 'played': False})
                pts  = calculate_playoff_match_points(rm, pred)
                entry['participants'][name] = {
                    'pred_g_local':     pred['g_local'],
                    'pred_g_visitante': pred['g_visitante'],
                    **pts,
                }
            playoff_all_matches.append(entry)

        # Calcular para cada participante
        participants_data = {}
        for name in PARTICIPANTS:
            ws_pred = wb[f'Predicciones_{name}']
            predictions = read_group_predictions(ws_pred)

            ws_playoffs = wb[f'Playoffs_{name}']
            pred_playoffs = read_playoffs(ws_playoffs)

            # Nueva hoja de predicciones detalladas de playoffs (Playoffs_Predicciones_*)
            ws_playoffs_det_name = f'Playoffs_Predicciones_{name}'
            pred_playoffs_detailed = None
            if ws_playoffs_det_name in wb.sheetnames:
                pred_playoffs_detailed = read_playoffs(wb[ws_playoffs_det_name])

            group_details = calculate_group_points(predictions, real_matches)
            playoff_details = calculate_playoff_points(pred_playoffs, real_playoffs)

            group_total = sum(m['total'] for m in group_details)
            playoff_total = playoff_details['total']
            
            # Sum up points from predicting the score of playoff matches
            participant_pm_index = playoff_match_preds[name]
            playoff_match_details = []
            playoff_match_total = 0
            for rm in real_playoff_matches:
                ronda = rm['ronda']
                pid   = rm['partido']
                pred  = participant_pm_index.get((ronda, pid), {'g_local': None, 'g_visitante': None, 'played': False})
                pts   = calculate_playoff_match_points(rm, pred)
                playoff_match_total += pts['total']
                playoff_match_details.append({
                    'partido':          pid,
                    'ronda':            ronda,
                    'local':            rm['local'],
                    'visitante':        rm['visitante'],
                    'real_g_local':     rm['g_local'],
                    'real_g_visitante': rm['g_visitante'],
                    'pred_g_local':     pred['g_local'],
                    'pred_g_visitante': pred['g_visitante'],
                    'played':           rm['played'],
                    **pts,
                })

            grand_total = group_total + playoff_total + playoff_match_total

            participants_data[name] = {
                'group_matches': group_details,
                'group_total': group_total,
                'playoffs': playoff_details,
                'playoff_total': playoff_total,
                'grand_total': grand_total,
                'playoff_detailed_predictions': pred_playoffs_detailed,
                'playoff_match_details': playoff_match_details,
            }

        # Rankings
        ranking = sorted(
            [{'name': n, 'total': participants_data[n]['grand_total']} for n in PARTICIPANTS],
            key=lambda x: x['total'],
            reverse=True
        )

        # Estadísticas de progreso del torneo
        total_group_matches = len(real_matches)  # 72
        played_group_matches = sum(1 for m in real_matches if m['g_local'] is not None and m['g_visitante'] is not None)
        groups_complete = sum(1 for g in standings.values() if g['complete'])
        groups_started = sum(1 for g in standings.values() if g['started'])

        # Puntos máximos posibles por participante (para rondas no jugadas)
        tournament_progress = {
            'total_group_matches': 72,
            'played_group_matches': played_group_matches,
            'pending_group_matches': 72 - played_group_matches,
            'groups_complete': groups_complete,
            'groups_started': groups_started,
            'groups_total': 12,
            'group_stage_complete': groups_complete == 12,
            'real_playoffs_rounds': {
                ronda: len(real_playoffs.get(ronda, [])) for ronda in PLAYOFF_COLUMNS
            }
        }

        # Predicciones futuras y predicciones completas de playoffs
        for name in PARTICIPANTS:
            ws_playoffs = wb[f'Playoffs_{name}']
            pred_playoffs = read_playoffs(ws_playoffs)

            # Guardamos TODAS las predicciones de playoff (para el bracket)
            participants_data[name]['playoff_predictions'] = {
                ronda: sorted(pred_playoffs.get(ronda, []))
                for ronda in PLAYOFF_COLUMNS
            }

            future_preds = {}
            for ronda in PLAYOFF_COLUMNS:
                real_teams_norm = [normalize_name(t) for t in real_playoffs.get(ronda, [])]
                pred_teams = pred_playoffs.get(ronda, [])
                
                is_round_complete = False
                if ronda == 'Octavos' and len(real_teams_norm) >= 16: is_round_complete = True
                elif ronda == 'Cuartos' and len(real_teams_norm) >= 8: is_round_complete = True
                elif ronda == 'Semifinal' and len(real_teams_norm) >= 4: is_round_complete = True
                elif ronda == 'Final' and len(real_teams_norm) >= 2: is_round_complete = True
                elif ronda == 'Campeón' and len(real_teams_norm) >= 1: is_round_complete = True
                elif ronda == 'Tercero' and len(real_teams_norm) >= 2: is_round_complete = True
                
                if not is_round_complete:
                    pending = []
                    for t in pred_teams:
                        t_n = normalize_name(t)
                        if t_n not in real_teams_norm and not is_eliminated(t_n, ronda):
                            pending.append(t)
                    if pending:
                        future_preds[ronda] = sorted(pending)
            participants_data[name]['future_predictions'] = future_preds

            # Puntos para predicciones detalladas si la hoja existe
            if participants_data[name].get('playoff_detailed_predictions') is not None:
                det = participants_data[name]['playoff_detailed_predictions']
                participants_data[name]['playoff_detailed_predictions'] = {
                    ronda: sorted(det.get(ronda, []))
                    for ronda in PLAYOFF_COLUMNS
                }

            # Puntos máximos posibles en grupos (partidos pendientes × 4pts max)
            pending_matches_pts = sum(
                4 for m in participants_data[name]['group_matches'] if not m['played']
            )
            # Puntos máximos posibles en quiniela de playoffs (partidos pendientes × 4pts max)
            pending_playoff_match_pts = sum(
                4 for m in participants_data[name]['playoff_match_details'] if not m['played']
            )
            
            # Puntos máximos posibles en playoffs por avance de equipos
            future_playoff_pts = 0
            for ronda in PLAYOFF_COLUMNS:
                teams = participants_data[name]['future_predictions'].get(ronda, [])
                pts_for_this_round = len(teams) * PLAYOFF_POINTS[ronda]
                
                # Deducción por colisiones: si predijo 2 equipos para la Final y se enfrentan en Semifinales, uno no llegará
                if ronda == 'Final' and len(teams) >= 2:
                    for rm in real_playoff_matches:
                        if rm['ronda'] == 'Semifinal':
                            local = normalize_name(rm['local'])
                            visitante = normalize_name(rm['visitante'])
                            # Si ambos equipos de la semifinal están en las predicciones de la final del usuario
                            teams_norm = [normalize_name(t) for t in teams]
                            if local in teams_norm and visitante in teams_norm:
                                pts_for_this_round -= PLAYOFF_POINTS['Final'] # Restamos los puntos de uno que seguro se elimina
                
                future_playoff_pts += pts_for_this_round

            participants_data[name]['max_possible_extra'] = pending_matches_pts + pending_playoff_match_pts + future_playoff_pts
            participants_data[name]['max_total'] = participants_data[name]['grand_total'] + participants_data[name]['max_possible_extra']

        return jsonify({
            'participants': participants_data,
            'ranking': ranking,
            'real_matches': real_matches,
            'real_playoffs': real_playoffs,
            'real_playoff_matches': real_playoff_matches,
            'playoff_all_matches': playoff_all_matches,
            'standings': standings,
            'best_8_thirds': best_8,
            'rest_4_thirds': rest_4,
            'all_thirds': all_thirds,
            'qualified': qualified,
            'tournament_progress': tournament_progress,
            'eliminated_from_tournament': list(eliminated_from_tournament),
            'eliminated_from_final': list(eliminated_from_final),
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


@app.route('/api/sync-playoffs', methods=['POST'])
def sync_playoffs():
    """
    Calcula los equipos clasificados al dieciseisavos y los escribe
    en la hoja 'Realidad Playoffs' del Excel (columna A = Dieciseisavos).
    """
    try:
        wb = load_workbook()
        ws_realidad = wb['Realidad']
        real_matches = read_group_predictions(ws_realidad)

        standings = calculate_group_standings(real_matches)
        best_8, rest_4, all_thirds = get_best_thirds(standings)
        qualified = get_qualified_teams(standings, best_8)

        if not qualified:
            return jsonify({'success': False, 'message': 'No hay equipos clasificados aún (no hay resultados de grupos).'})

        written = write_qualified_to_excel(qualified, best_8)
        return jsonify({
            'success': written,
            'message': f'Se escribieron {len(qualified)} equipos en Realidad Playoffs.',
            'qualified': [t['equipo'] for t in qualified],
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


@app.route('/api/sync-and-reload', methods=['POST'])
def sync_and_reload():
    """
    Endpoint que en un solo call:
    1. Calcula clasificados de grupos
    2. Escribe en Realidad Playoffs (col A)
    3. Retorna los datos completos actualizados (como /api/data)
    """
    try:
        wb = load_workbook()
        ws_realidad = wb['Realidad']
        real_matches = read_group_predictions(ws_realidad)

        standings = calculate_group_standings(real_matches)
        best_8, rest_4, all_thirds = get_best_thirds(standings)
        qualified = get_qualified_teams(standings, best_8)

        sync_msg = 'Sin equipos para guardar aún.'
        if qualified:
            write_qualified_to_excel(qualified, best_8)
            sync_msg = f'{len(qualified)} clasificados guardados en Realidad Playoffs.'

        # Usar la misma logica de get_data
        res = get_data()
        if res.status_code == 200:
            data = res.get_json()
            data['sync_message'] = sync_msg
            return jsonify(data)
        else:
            return res

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


def _pause_before_exit():
    """Mantiene la consola abierta en el .exe para que se lean errores."""
    if getattr(sys, 'frozen', False):
        input('\nPresiona Enter para cerrar...')


if __name__ == '__main__':
    IS_FROZEN = getattr(sys, 'frozen', False)
    PORT = 5001

    try:
        if IS_FROZEN and not os.path.exists(EXCEL_PATH):
            print('\nERROR: No se encontró el archivo Excel.')
            print(f'  Ruta esperada: {EXCEL_PATH}')
            print('\nColoca "Polla Mundial 2026.xlsx" en la misma carpeta que el .exe.')
            _pause_before_exit()
            sys.exit(1)

        if IS_FROZEN:
            def open_browser():
                webbrowser.open(f'http://127.0.0.1:{PORT}')
            threading.Timer(1.5, open_browser).start()
            print(f'\n⚽ Polla Mundial 2026 corriendo en http://127.0.0.1:{PORT}')
            print('   Cierra esta ventana para detener la app.\n')
            app.run(debug=False, port=PORT, use_reloader=False)
        else:
            app.run(debug=True, port=PORT)
    except Exception as e:
        import traceback
        print(f'\nERROR al iniciar la app: {e}')
        traceback.print_exc()
        _pause_before_exit()
        sys.exit(1)
